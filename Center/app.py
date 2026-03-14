"""
Center Location Dashboard - Flask Web Application
Data is loaded ONCE at startup into memory — all requests serve from RAM.
"""
from flask import Flask, render_template, jsonify, request
import openpyxl, os, time

app = Flask(__name__)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR  = os.path.join(BASE_DIR, "data")

TOC_COLORS = {
    "Center":              "#10b981",
    "Community":           "#3b82f6",
    "Hospital / UPHC/PHC": "#ef4444",
    "School and College":  "#f97316",
}
TOC_DEFAULT = "#a855f7"

# ── Load ALL excel data once at module import (startup) ──────────────────────
def _read_sheet(path, sheet, skip_col=0):
    if not os.path.exists(path):
        print(f"  WARNING: {path} not found")
        return []
    t = time.time()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    raw_h = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    headers = [str(h).strip() if h else "" for h in raw_h]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[skip_col]:
            continue
        r = {}
        for i, h in enumerate(headers):
            v = row[i] if i < len(row) else None
            r[h] = float(v) if isinstance(v, (int, float)) else (str(v).strip() if v else "")
        rows.append(r)
    wb.close()
    print(f"  Loaded {os.path.basename(path)}[{sheet}]: {len(rows)} rows in {time.time()-t:.2f}s")
    if rows:
        print(f"  Columns in {os.path.basename(path)}: {list(rows[0].keys())}")
    return rows

print("Loading data files...")
CENTERS   = _read_sheet(os.path.join(DATA_DIR, "MASTER_NEW.xlsx"),      "Sheet1",      skip_col=1)
IMPACT    = _read_sheet(os.path.join(DATA_DIR, "Impact.xlsx"),           "Sheet1",      skip_col=0)
CLUSTERS  = _read_sheet(os.path.join(DATA_DIR, "Cluster_Master.xlsx"),  "Sheet1",      skip_col=0)
CRITERIA  = _read_sheet(os.path.join(DATA_DIR, "Criteria.xlsx"),        "Center Wise", skip_col=0)
EMPLOYER  = _read_sheet(os.path.join(DATA_DIR, "Employer_Master.xlsx"), "Sheet1",      skip_col=0)
MANPOWER  = _read_sheet(os.path.join(DATA_DIR, "Manpower.xlsx"),        "Sheet1",      skip_col=0)
MIGRANT   = _read_sheet(os.path.join(DATA_DIR, "migrant_Data.xlsx"),    "Sheet1",      skip_col=0)

# Pre-build cluster → map URL lookup
CLUSTER_MAP_URL = {
    c.get("Sub Cluster Name", "").strip(): c.get("Map", "")
    for c in CLUSTERS
}

# Pre-build Center ID → MASTER row lookup (for Manpower enrichment)
CENTER_ID_MAP = {}
for c in CENTERS:
    raw = c.get("Centre ID", "")
    cid = str(int(raw) if isinstance(raw, float) and raw == int(raw) else raw).strip() if raw else ""
    if cid:
        CENTER_ID_MAP[cid] = c

print(f"Ready. {len(CENTERS)} centers | {len(IMPACT)} impact | {len(CRITERIA)} criteria | "
      f"{len(EMPLOYER)} employer | {len(MANPOWER)} manpower | {len(MIGRANT)} migrant rows")


# ── tiny helpers ──────────────────────────────────────────────────────────────
def _set(items, key):
    return sorted(set(r.get(key, "") for r in items if r.get(key, "")))

def _filt(pool, **kwargs):
    """Filter list of dicts; skip empty filter values."""
    for key, val in kwargs.items():
        if val:
            pool = [r for r in pool if r.get(key, "") == val]
    return pool

def _find_col(row, *candidates):
    for c in candidates:
        if c in row:
            return row[c]
    row_lower = {k.strip().lower(): v for k, v in row.items()}
    for c in candidates:
        if c.strip().lower() in row_lower:
            return row_lower[c.strip().lower()]
    return 0


# ── ROUTES ────────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/debug/impact-cols")
def debug_impact_cols():
    if IMPACT:
        return jsonify({"columns": list(IMPACT[0].keys()), "sample_row": IMPACT[0]})
    return jsonify({"columns": [], "sample_row": {}})


# ── Cascading filter endpoints ────────────────────────────────────────────────

@app.route("/api/filters/active-centers")
def ac_filters():
    g = request.args.get
    sel_cluster  = g("cluster","");  sel_sub    = g("sub_cluster","")
    sel_toc      = g("toc","");      sel_status = g("status","")
    sel_dist     = g("dist_slab",""); sel_entity = g("entity","")

    cl_pool  = _filt(CENTERS, **{"Cluster Name":     sel_cluster})
    sub_pool = _filt(cl_pool,  **{"Sub Cluster Name": sel_sub})
    t_pool   = _filt(sub_pool, **{"Entity ": sel_entity, "Type of Center": sel_toc,
                                   "Status": sel_status, "Distance Slab":  sel_dist})

    return jsonify({
        "clusters":     _set(CENTERS,  "Cluster Name"),
        "sub_clusters": _set(cl_pool,  "Sub Cluster Name"),
        "entities":     _set(sub_pool, "Entity "),
        "tocs":         _set(sub_pool, "Type of Center"),
        "statuses":     _set(CENTERS,  "Status"),
        "dist_slabs":   _set(CENTERS,  "Distance Slab"),
        "centers_near": _set(t_pool,   "Centre Name  Near"),
        "toc_colors":   TOC_COLORS,
    })


@app.route("/api/filters/qp-criteria")
def qp_filters():
    g = request.args.get
    sel_cluster = g("cluster",""); sel_sub    = g("sub_cluster","")
    sel_sector  = g("sector","");  sel_bv     = g("bv","")
    sel_entity  = g("entity","")

    cl_pool  = _filt(CRITERIA, **{"Cluster Name":     sel_cluster})
    sub_pool = _filt(cl_pool,  **{"Sub Cluster Name": sel_sub})
    c_pool   = _filt(sub_pool, **{"Sector_name": sel_sector,
                                   "Business Vertical ": sel_bv,
                                   "Entity ": sel_entity})
    return jsonify({
        "clusters":           _set(CRITERIA,  "Cluster Name"),
        "sub_clusters":       _set(cl_pool,   "Sub Cluster Name"),
        "entities":           _set(sub_pool,  "Entity "),
        "sectors":            _set(sub_pool,  "Sector_name"),
        "business_verticals": _set(sub_pool,  "Business Vertical "),
        "centers":            _set(c_pool,    "Centre Name"),
    })


@app.route("/api/filters/impact")
def impact_filters():
    g = request.args.get
    sel_cluster  = g("cluster","");  sel_sub   = g("sub_cluster","")
    sel_ctype    = g("centre_type","")

    cl_pool   = _filt(IMPACT, **{"Cluster Name":     sel_cluster})
    sub_pool  = _filt(cl_pool, **{"Sub Cluster Name": sel_sub})
    ct_pool   = _filt(sub_pool, **{"Centre Type":     sel_ctype})

    return jsonify({
        "clusters":      _set(IMPACT,    "Cluster Name"),
        "sub_clusters":  _set(cl_pool,   "Sub Cluster Name"),
        "centre_types":  _set(sub_pool,  "Centre Type"),
        "centers":       _set(ct_pool,   "Centre Name"),
    })


@app.route("/api/filters/employer")
def employer_filters():
    g = request.args.get
    sel_cluster = g("cluster",""); sel_sub = g("sub_cluster","")
    sel_center  = g("center","");  sel_qp  = g("qp","")

    cl_pool  = _filt(EMPLOYER, **{"Cluster Name":     sel_cluster})
    sub_pool = _filt(cl_pool,  **{"Sub Cluster Name": sel_sub})
    c_pool   = _filt(sub_pool, **{"Centre Name":      sel_center})

    return jsonify({
        "clusters":     _set(EMPLOYER, "Cluster Name"),
        "sub_clusters": _set(cl_pool,  "Sub Cluster Name"),
        "centers":      _set(sub_pool, "Centre Name"),
        "qps":          _set(c_pool,   "QP Name"),
    })


@app.route("/api/filters/migrant")
def migrant_filters():
    g = request.args.get
    sel_cluster = g("cluster",""); sel_sub    = g("sub_cluster","")
    sel_center  = g("center","");  sel_sector = g("sector","")
    sel_fy      = g("fy","")

    cl_pool  = _filt(MIGRANT, **{"Cluster Name":     sel_cluster})
    sub_pool = _filt(cl_pool,  **{"Sub Cluster Name": sel_sub})
    c_pool   = _filt(sub_pool, **{"Centre Name":      sel_center})
    s_pool   = _filt(c_pool,   **{"Sector_name":      sel_sector})

    return jsonify({
        "clusters":     _set(MIGRANT,  "Cluster Name"),
        "sub_clusters": _set(cl_pool,  "Sub Cluster Name"),
        "centers":      _set(sub_pool, "Centre Name"),
        "sectors":      _set(sub_pool, "Sector_name"),
        "fys":          _set(MIGRANT,  "Project Executed FY"),
        "qps":          _set(s_pool,   "QP Name"),
    })


# ── Data endpoints ────────────────────────────────────────────────────────────

@app.route("/api/data/active-centers")
def api_active_centers():
    g = request.args.get
    filtered = _filt(CENTERS,
        **{"Entity ":           g("entity",""),
           "Cluster Name":      g("cluster",""),
           "Sub Cluster Name":  g("sub_cluster",""),
           "Type of Center":    g("toc",""),
           "Status":            g("status",""),
           "Distance Slab":     g("dist_slab",""),
           "Centre Name  Near": g("center_near","")})

    toc_counts = {}
    for r in filtered:
        t = r.get("Type of Center","Unknown")
        toc_counts[t] = toc_counts.get(t,0) + 1

    map_points, seen = [], set()
    for r in filtered:
        name = r.get("Centre Name","")
        lat  = r.get("LAT",0); lng = r.get("LAG",0)
        if name not in seen and lat and lng:
            seen.add(name)
            sub = r.get("Sub Cluster Name","")
            toc = r.get("Type of Center","")
            raw_cid = r.get("Centre ID","")
            cid = str(int(raw_cid) if isinstance(raw_cid, float) and raw_cid == int(raw_cid) else raw_cid).strip() if raw_cid else ""

            # Attach manpower for this center
            mp_rows = [m for m in MANPOWER if str(m.get("Center ID","")).strip() == cid]
            manpower_list = [{"name": m.get("Employee Name",""), "role": m.get("Role","")}
                             for m in mp_rows]

            map_points.append({
                "name":        name,
                "near":        r.get("Centre Name  Near",""),
                "lat": lat, "lng": lng,
                "cluster":     r.get("Cluster Name",""),
                "sub_cluster": sub,
                "toc":         toc,
                "color":       TOC_COLORS.get(toc, TOC_DEFAULT),
                "status":      r.get("Status",""),
                "address":     r.get("Center address",""),
                "distance":    r.get("Distance (km)",0),
                "dist_slab":   r.get("Distance Slab",""),
                "map_url":     CLUSTER_MAP_URL.get(sub,""),
                "manpower":    manpower_list,
            })

    return jsonify({
        "summary":    {"toc_counts": toc_counts, "total_records": len(filtered)},
        "map_points": map_points,
        "toc_colors": TOC_COLORS,
        "table":      filtered[:1000]
    })


@app.route("/api/data/qp-criteria")
def api_qp_criteria():
    g = request.args.get
    filtered = _filt(CRITERIA,
        **{"Entity ":            g("entity",""),
           "Cluster Name":       g("cluster",""),
           "Sub Cluster Name":   g("sub_cluster",""),
           "Centre Name":        g("center_name",""),
           "Sector_name":        g("sector",""),
           "Business Vertical ": g("bv","")})
    return jsonify({"data": filtered[:1000], "total": len(filtered)})


@app.route("/api/data/impact")
def api_impact():
    g = request.args.get
    filtered = _filt(IMPACT,
        **{"Cluster Name":     g("cluster",""),
           "Sub Cluster Name": g("sub_cluster",""),
           "Centre Type":      g("centre_type",""),
           "Centre Name":      g("center_name","")})

    def n(v):
        try: return int(float(v or 0))
        except: return 0

    cluster_summary = {}
    for r in filtered:
        cl = r.get("Cluster Name","Other")
        if cl not in cluster_summary:
            cluster_summary[cl] = {"enrolled":0,"certified":0,"placed":0,"dropout":0}
        cluster_summary[cl]["enrolled"]  += n(_find_col(r, "Enrolled", " Enrolled", "enrolled"))
        cluster_summary[cl]["certified"] += n(_find_col(r, "Certified", " Certified", "certified"))
        cluster_summary[cl]["placed"]    += n(_find_col(r, "Placed", " Placed", "placed"))
        cluster_summary[cl]["dropout"]   += n(_find_col(r, "DropOut/Notcertified", " DropOut/Notcertified",
                                                            "Dropout", " Dropout", "dropout"))

    return jsonify({
        "summary": {k: sum(v[k] for v in cluster_summary.values())
                    for k in ("enrolled","certified","placed","dropout")},
        "cluster_summary": cluster_summary,
        "data": filtered[:1000]
    })


@app.route("/api/data/employer")
def api_employer():
    g = request.args.get
    filtered = _filt(EMPLOYER,
        **{"Cluster Name":      g("cluster",""),
           "Sub Cluster Name":  g("sub_cluster",""),
           "Centre Name":       g("center",""),
           "QP Name":           g("qp","")})

    def n(v):
        try: return int(float(v or 0))
        except: return 0

    employer_totals = {}
    for r in filtered:
        emp = r.get("Employer Name","Unknown")
        employer_totals[emp] = employer_totals.get(emp, 0) + n(r.get("Placed Count", 0))

    total_placed    = sum(employer_totals.values())
    total_employers = len(employer_totals)
    total_qps       = len(set(r.get("QP Name","") for r in filtered if r.get("QP Name","")))
    sorted_employers = sorted(employer_totals.items(), key=lambda x: -x[1])

    return jsonify({
        "summary": {
            "total_placed":    total_placed,
            "total_employers": total_employers,
            "total_qps":       total_qps,
            "total_records":   len(filtered),
        },
        "employer_totals": sorted_employers,
        "data": filtered[:2000],
    })


@app.route("/api/filters/manpower")
def manpower_filters():
    g = request.args.get
    sel_cluster = g("cluster",""); sel_sub     = g("sub_cluster","")
    sel_project = g("project",""); sel_center  = g("center","")

    # Enrich all manpower rows with master data first
    enriched = []
    for m in MANPOWER:
        raw_cid = m.get("Center ID","")
        cid = str(int(raw_cid) if isinstance(raw_cid, float) and raw_cid == int(raw_cid) else raw_cid).strip() if raw_cid else ""
        master = CENTER_ID_MAP.get(cid, {})
        enriched.append({
            "_cluster": master.get("Cluster Name",""),
            "_sub":     master.get("Sub Cluster Name",""),
            "_project": master.get("Project",""),
            "_center":  master.get("Centre Name",""),
        })

    all_rows    = enriched
    cl_pool     = [r for r in all_rows if not sel_cluster or r["_cluster"] == sel_cluster]
    sub_pool    = [r for r in cl_pool  if not sel_sub     or r["_sub"]     == sel_sub]
    proj_pool   = [r for r in sub_pool if not sel_project or r["_project"] == sel_project]

    return jsonify({
        "clusters":     sorted(set(r["_cluster"] for r in all_rows  if r["_cluster"])),
        "sub_clusters": sorted(set(r["_sub"]     for r in cl_pool   if r["_sub"])),
        "projects":     sorted(set(r["_project"] for r in sub_pool  if r["_project"])),
        "centers":      sorted(set(r["_center"]  for r in proj_pool if r["_center"])),
    })


@app.route("/api/data/manpower")
def api_manpower():
    g = request.args.get
    sel_project = g("project",""); sel_cluster = g("cluster","")
    sel_sub     = g("sub_cluster",""); sel_center = g("center","")

    data = []
    for m in MANPOWER:
        raw_cid = m.get("Center ID","")
        cid     = str(int(raw_cid) if isinstance(raw_cid, float) and raw_cid == int(raw_cid) else raw_cid).strip() if raw_cid else ""
        master  = CENTER_ID_MAP.get(cid, {})
        project    = master.get("Project","")
        cluster    = master.get("Cluster Name","")
        sub        = master.get("Sub Cluster Name","")
        centre     = master.get("Centre Name","")

        if sel_project and project != sel_project: continue
        if sel_cluster and cluster != sel_cluster: continue
        if sel_sub     and sub     != sel_sub:     continue
        if sel_center  and centre  != sel_center:  continue

        data.append({
            "sub_project_code": m.get("Sub Project Code",""),
            "center_id":        cid,
            "employee_name":    m.get("Employee Name",""),
            "role":             m.get("Role",""),
            "project":          project,
            "centre_name":      centre,
            "cluster":          cluster,
            "sub_cluster":      sub,
            "state":            master.get("State",""),
            "district":         master.get("District",""),
            "status":           master.get("Status",""),
            "toc":              master.get("Type of Center",""),
            "lat":              master.get("LAT",0),
            "lng":              master.get("LAG",""),
        })

    role_counts = {}
    for d in data:
        r = d["role"] or "Unknown"
        role_counts[r] = role_counts.get(r,0) + 1

    return jsonify({
        "data":        data,
        "total":       len(data),
        "role_counts": role_counts,
    })


@app.route("/api/data/migrant")
def api_migrant():
    g = request.args.get
    filtered = _filt(MIGRANT,
        **{"Cluster Name":        g("cluster",""),
           "Sub Cluster Name":    g("sub_cluster",""),
           "Centre Name":         g("center",""),
           "Sector_name":         g("sector",""),
           "Project Executed FY": g("fy",""),
           "QP Name":             g("qp","")})

    def n(v):
        try: return int(float(v or 0))
        except: return 0

    center_agg = {}
    area_agg   = {}

    for r in filtered:
        cname   = r.get("Centre Name","Unknown")
        area    = r.get("Area","Unknown")
        cluster = r.get("Cluster Name","")
        sub     = r.get("Sub Cluster Name","")

        enrolled   = n(_find_col(r, "Total  Enrolled", "Total Enrolled"))
        interstate = n(_find_col(r, "Inter state migrant", "  Inter state migrant"))
        interdistr = n(_find_col(r, "Inter district Migrants", "  Inter district Migrants"))
        local      = n(_find_col(r, "Local", "  Local"))
        community  = n(_find_col(r, "Count Community", "  Count Community "))

        for agg, key in [(center_agg, cname), (area_agg, area)]:
            if key not in agg:
                agg[key] = {"cluster": cluster, "sub_cluster": sub,
                            "enrolled":0,"interstate":0,"interdistrict":0,
                            "local":0,"community":0}
            agg[key]["enrolled"]      += enrolled
            agg[key]["interstate"]    += interstate
            agg[key]["interdistrict"] += interdistr
            agg[key]["local"]         += local
            agg[key]["community"]     += community

    return jsonify({
        "summary": {
            "enrolled":      sum(v["enrolled"]      for v in center_agg.values()),
            "interstate":    sum(v["interstate"]    for v in center_agg.values()),
            "interdistrict": sum(v["interdistrict"] for v in center_agg.values()),
            "local":         sum(v["local"]         for v in center_agg.values()),
            "community":     sum(v["community"]     for v in center_agg.values()),
        },
        "center_agg": center_agg,
        "area_agg":   area_agg,
        "data":       filtered[:2000],
    })


if __name__ == "__main__":
    app.run(debug=True, port=5000)
